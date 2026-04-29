from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .normalizers import clean_text

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
READY_FILL = PatternFill("solid", fgColor="C6EFCE")
REVIEW_FILL = PatternFill("solid", fgColor="FFC7CE")
WARNING_FILL = PatternFill("solid", fgColor="FFEB9C")
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def read_sheet(path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [clean_text(h) for h in next(rows)]
    except StopIteration:
        return []
    # Skip generated title rows if this workbook structure is used.
    if len([h for h in headers if h]) <= 2:
        for _ in range(2):
            try:
                headers = [clean_text(h) for h in next(rows)]
            except StopIteration:
                return []
    result: list[dict[str, Any]] = []
    for row in rows:
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]}
        if any(clean_text(v) for v in item.values()):
            result.append(item)
    return result

def read_reference_sheet(path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read sheets from the generated reference workbook where headers are on row 3."""
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    header_row = 3
    headers = [clean_text(cell.value) for cell in ws[header_row]]
    data: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]}
        if any(clean_text(v) for v in item.values()):
            data.append(item)
    return data

def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(clean_text(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

def style_header(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def write_workbook(path: str | Path, output_rows: list[dict[str, Any]], validation_rows: list[dict[str, Any]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Magento_Import"
    _write_rows(ws, output_rows)
    _add_table(ws, "MagentoImportTable")
    _status_fill(ws, "validation_status")

    vr = wb.create_sheet("Validation_Report")
    _write_rows(vr, validation_rows)
    _add_table(vr, "ValidationReportTable")
    _status_fill(vr, "severity")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

def _write_rows(ws, rows: list[dict[str, Any]]):
    if not rows:
        ws.append(["message"])
        ws.append(["No rows generated"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_header(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    autosize(ws)

def _add_table(ws, name: str):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

def _status_fill(ws, field_name: str):
    headers = [clean_text(c.value) for c in ws[1]]
    if field_name not in headers:
        return
    idx = headers.index(field_name) + 1
    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
        cell = row[0]
        text = clean_text(cell.value).casefold()
        if "ready" in text or "info" in text:
            cell.fill = READY_FILL
        elif "warning" in text or "manual" in text:
            cell.fill = WARNING_FILL
        elif "error" in text or "blocker" in text:
            cell.fill = REVIEW_FILL
